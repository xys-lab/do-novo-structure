import csv
import re
from pathlib import Path
from array import array
import math

import numpy as np
from openpyxl import load_workbook
from tqdm import tqdm


# =========================================================
# 1. 配置区
# =========================================================
INPUT_FILES = [
    r"D:\实验\数据\DSSTox_Feb_2024\DSSToxDump1.xlsx",
     r"D:\实验\数据\DSSTox_Feb_2024\DSSToxDump2.xlsx",
     r"D:\实验\数据\DSSTox_Feb_2024\DSSToxDump3.xlsx",
r"D:\实验\数据\DSSTox_Feb_2024\DSSToxDump4.xlsx",
r"D:\实验\数据\DSSTox_Feb_2024\DSSToxDump5.xlsx",
r"D:\实验\数据\DSSTox_Feb_2024\DSSToxDump6.xlsx",
r"D:\实验\数据\DSSTox_Feb_2024\DSSToxDump7.xlsx",
r"D:\实验\数据\DSSTox_Feb_2024\DSSToxDump8.xlsx",
r"D:\实验\数据\DSSTox_Feb_2024\DSSToxDump9.xlsx",
r"D:\实验\数据\DSSTox_Feb_2024\DSSToxDump10.xlsx",
r"D:\实验\数据\DSSTox_Feb_2024\DSSToxDump11.xlsx",
r"D:\实验\数据\DSSTox_Feb_2024\DSSToxDump12.xlsx",
r"D:\实验\数据\DSSTox_Feb_2024\DSSToxDump13.xlsx",
]

SMILES_COL = "SMILES"

OUTPUT_DETAIL_FILE = r"D:\实验\数据\DSSTox_Feb_2024\smiles_length_detail_stream.csv"
OUTPUT_SUMMARY_FILE = r"D:\实验\数据\DSSTox_Feb_2024\smiles_length_summary_stream.csv"
OUTPUT_RECOMMEND_FILE = r"D:\实验\数据\DSSTox_Feb_2024\smiles_padding_recommendation.csv"


# =========================================================
# 2. 你的 tokenizer 规则（严格照你的版本）
# =========================================================
TOKEN_PATTERN = re.compile(
    r"\[[^\]]+\]"         # [NH4+], [C@H], [O-]
    r"|Br|Cl|Si|Na|Ca"    # 常见双字符原子
    r"|%\d{2}"            # %12 环编号
    r"|\d"                # 单个数字
    r"|\(|\)|\.|=|#|-|\+|\\|/|:|~|\?|>|<|\*|\$"
    r"|[A-Za-z]"          # 单字符原子
)

def tokenize_smiles(smiles: str):
    return TOKEN_PATTERN.findall(smiles)


# =========================================================
# 3. 工具函数
# =========================================================
def round_up_nice(x: float) -> int:
    """
    把长度向上取整到更适合模型设定的值。
    优先取常见长度：32, 48, 64, 80, 96, 112, 128, 160, 192, 224, 256, ...
    """
    candidates = [32, 48, 64, 80, 96, 112, 128, 160, 192, 224, 256, 320, 384, 448, 512, 640, 768, 896, 1024]
    for c in candidates:
        if x <= c:
            return c
    return int(math.ceil(x / 64.0) * 64)


def summarize_array(arr_vals, metric_name):
    if len(arr_vals) == 0:
        return {
            "metric": metric_name,
            "count": 0,
            "min": None,
            "mean": None,
            "median": None,
            "p90": None,
            "p95": None,
            "p99": None,
            "p99.5": None,
            "max": None,
        }

    vals = np.array(arr_vals, dtype=np.int32)
    return {
        "metric": metric_name,
        "count": int(len(vals)),
        "min": int(vals.min()),
        "mean": round(float(vals.mean()), 4),
        "median": round(float(np.median(vals)), 4),
        "p90": round(float(np.percentile(vals, 90)), 4),
        "p95": round(float(np.percentile(vals, 95)), 4),
        "p99": round(float(np.percentile(vals, 99)), 4),
        "p99.5": round(float(np.percentile(vals, 99.5)), 4),
        "max": int(vals.max()),
    }


def count_non_empty_rows_in_excel(file_path, smiles_col_name):
    """
    只用于 tqdm 总数显示。会扫描一遍 Excel。
    如果你不在乎进度条总数，可以不调用这个函数。
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header_clean = [str(x).strip() if x is not None else "" for x in header]

    if smiles_col_name not in header_clean:
        wb.close()
        raise ValueError(f"{file_path} 中找不到列 {smiles_col_name}，实际列为：{header_clean}")

    smiles_idx = header_clean.index(smiles_col_name)

    count = 0
    for row in rows:
        if row is None:
            continue
        val = row[smiles_idx] if smiles_idx < len(row) else None
        if val is None:
            continue
        s = str(val).strip()
        if s != "":
            count += 1

    wb.close()
    return count


def iter_smiles_from_excel(file_path, smiles_col_name):
    """
    流式读取 Excel 第一张表中的 SMILES 列。
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header_clean = [str(x).strip() if x is not None else "" for x in header]

    if smiles_col_name not in header_clean:
        wb.close()
        raise ValueError(f"{file_path} 中找不到列 {smiles_col_name}，实际列为：{header_clean}")

    smiles_idx = header_clean.index(smiles_col_name)

    for row in rows:
        if row is None:
            continue
        val = row[smiles_idx] if smiles_idx < len(row) else None
        if val is None:
            continue
        s = str(val).strip()
        if s == "":
            continue
        yield s

    wb.close()


# =========================================================
# 4. 主程序：流式统计
# =========================================================
def main():
    Path(OUTPUT_DETAIL_FILE).parent.mkdir(parents=True, exist_ok=True)

    # 用紧凑整数数组保存长度，内存比 list 更省
    raw_char_lens = array("I")
    raw_token_lens = array("I")
    raw_seq_lens = array("I")

    total_count = 0

    with open(OUTPUT_DETAIL_FILE, "w", newline="", encoding="utf-8-sig") as f_detail:
        writer = csv.DictWriter(
            f_detail,
            fieldnames=[
                "source_file",
                "row_index_in_source",
                "smiles_raw",
                "raw_char_len",
                "raw_token_len",
                "raw_seq_len_with_special",
            ]
        )
        writer.writeheader()

        for file_path in INPUT_FILES:
            print(f"\n正在处理文件: {file_path}")

            total_rows = count_non_empty_rows_in_excel(file_path, SMILES_COL)

            for idx, smi in enumerate(
                tqdm(
                    iter_smiles_from_excel(file_path, SMILES_COL),
                    total=total_rows,
                    desc=f"Processing {Path(file_path).name}"
                ),
                start=2  # Excel 第1行是表头，所以数据从第2行开始
            ):
                tokens = tokenize_smiles(smi)
                raw_char_len = len(smi)
                raw_token_len = len(tokens)
                raw_seq_len_with_special = raw_token_len + 2  # <SOS> + <EOS>

                writer.writerow({
                    "source_file": file_path,
                    "row_index_in_source": idx,
                    "smiles_raw": smi,
                    "raw_char_len": raw_char_len,
                    "raw_token_len": raw_token_len,
                    "raw_seq_len_with_special": raw_seq_len_with_special,
                })

                raw_char_lens.append(raw_char_len)
                raw_token_lens.append(raw_token_len)
                raw_seq_lens.append(raw_seq_len_with_special)

                total_count += 1

    print("\n详细长度结果已保存到:")
    print(OUTPUT_DETAIL_FILE)

    # =====================================================
    # 5. 生成 summary
    # =====================================================
    summary_rows = []
    summary_rows.append({
        "metric": "overall_count",
        "count": total_count,
        "min": None,
        "mean": None,
        "median": None,
        "p90": None,
        "p95": None,
        "p99": None,
        "p99.5": None,
        "max": None,
    })
    summary_rows.append(summarize_array(raw_char_lens, "raw_char_len"))
    summary_rows.append(summarize_array(raw_token_lens, "raw_token_len"))
    summary_rows.append(summarize_array(raw_seq_lens, "raw_seq_len_with_special"))

    with open(OUTPUT_SUMMARY_FILE, "w", newline="", encoding="utf-8-sig") as f_sum:
        writer = csv.DictWriter(
            f_sum,
            fieldnames=["metric", "count", "min", "mean", "median", "p90", "p95", "p99", "p99.5", "max"]
        )
        writer.writeheader()
        writer.writerows(summary_rows)

    print("摘要结果已保存到:")
    print(OUTPUT_SUMMARY_FILE)

    # =====================================================
    # 6. 自动给出 padding 推荐
    #    真正最重要的是 raw_seq_len_with_special
    # =====================================================
    seq_summary = summarize_array(raw_seq_lens, "raw_seq_len_with_special")

    p95 = seq_summary["p95"]
    p99 = seq_summary["p99"]
    p995 = seq_summary["p99.5"]
    max_len = seq_summary["max"]

    recommend_rows = [
        {
            "strategy": "conservative_efficiency",
            "based_on": "p95",
            "observed_value": p95,
            "recommended_padding_len": round_up_nice(p95),
            "comment": "更省显存/更高训练效率，但会丢掉更多长尾样本"
        },
        {
            "strategy": "balanced_default",
            "based_on": "p99",
            "observed_value": p99,
            "recommended_padding_len": round_up_nice(p99),
            "comment": "最推荐的第一版训练长度，通常是效率和覆盖率的平衡点"
        },
        {
            "strategy": "high_coverage",
            "based_on": "p99.5",
            "observed_value": p995,
            "recommended_padding_len": round_up_nice(p995),
            "comment": "覆盖更多长尾样本，但训练更慢、padding浪费更大"
        },
        {
            "strategy": "do_not_use_directly",
            "based_on": "max",
            "observed_value": max_len,
            "recommended_padding_len": round_up_nice(max_len),
            "comment": "最大值通常受极少数异常超长样本影响，不建议直接拿来设padding"
        },
    ]

    with open(OUTPUT_RECOMMEND_FILE, "w", newline="", encoding="utf-8-sig") as f_rec:
        writer = csv.DictWriter(
            f_rec,
            fieldnames=["strategy", "based_on", "observed_value", "recommended_padding_len", "comment"]
        )
        writer.writeheader()
        writer.writerows(recommend_rows)

    print("padding 建议已保存到:")
    print(OUTPUT_RECOMMEND_FILE)

    # =====================================================
    # 7. 终端打印关键结果
    # =====================================================
    print("\n==============================")
    print("Length Summary")
    print("==============================")
    for row in summary_rows:
        if row["metric"] == "overall_count":
            print(f"总条数: {row['count']}")
        else:
            print(
                f"{row['metric']:28s} "
                f"count={row['count']}, "
                f"p95={row['p95']}, "
                f"p99={row['p99']}, "
                f"p99.5={row['p99.5']}, "
                f"max={row['max']}"
            )

    print("\n==============================")
    print("Padding Recommendation")
    print("==============================")
    for r in recommend_rows:
        print(
            f"{r['strategy']:22s} "
            f"{r['based_on']:6s}={r['observed_value']}  "
            f"-> recommended max_len = {r['recommended_padding_len']}"
        )


if __name__ == "__main__":
    main()