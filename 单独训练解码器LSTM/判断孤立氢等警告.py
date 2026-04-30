import csv
from pathlib import Path
from openpyxl import load_workbook
from tqdm import tqdm

from rdkit import Chem, RDLogger
from rdkit.Chem import Descriptors, rdMolDescriptors

# =========================================================
# 配置区
# =========================================================
INPUT_FILES = [
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump1_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump2_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump3_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump4_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump5_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump6_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump7_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump8_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump9_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump10_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump11_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump12_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump13_filled.xlsx",
   # 可以继续添加更多文件
]

SMILES_COL = "SMILES"
INPUT_INCHIKEY_COL = "INCHIKEY"
MONOISOTOPIC_MASS_COL = "MONOISOTOPIC_MASS"

MAX_EXACT_MW = 1500.0
FILTER_ATOMS = {'C', 'N', 'O', 'S', 'F', 'Cl', 'Br', 'I', 'P', 'H'}

OUTPUT_DIR = Path(r"D:\实验\数据\DSSTox_Feb_2024\strict_cleaning_5cols")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

CLEAN_FILE = OUTPUT_DIR / "cleaned_DSSTox_5cols.csv"
ISSUE_FILE = OUTPUT_DIR / "problem_records_DSSTox_5cols.csv"

# =========================================================
# 静默 RDKit warning
# =========================================================
RDLogger.logger().setLevel(RDLogger.CRITICAL)


# =========================================================
# Excel 流式读取（修复表头误读）
# =========================================================
def iter_rows_from_excel(file_path, smiles_col, inchikey_col, mono_mass_col):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[ws.sheetnames[0]] if False else wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header_clean = [str(x).strip() if x is not None else "" for x in header]

    required_cols = [smiles_col, inchikey_col, mono_mass_col]
    for col in required_cols:
        if col not in header_clean:
            wb.close()
            raise ValueError(f"{file_path} 中找不到必要列: {col}")

    smiles_idx = header_clean.index(smiles_col)
    inchikey_idx = header_clean.index(inchikey_col)
    mono_mass_idx = header_clean.index(mono_mass_col)

    for row_idx, row in enumerate(rows, start=2):
        if row is None:
            continue

        smi = row[smiles_idx] if smiles_idx < len(row) else None
        input_ik = row[inchikey_idx] if inchikey_idx < len(row) else None
        mono_mass = row[mono_mass_idx] if mono_mass_idx < len(row) else None

        smi = str(smi).strip() if smi is not None else None
        input_ik = str(input_ik).strip() if input_ik is not None else None

        yield row_idx, smi, input_ik, mono_mass

    wb.close()


# =========================================================
# 单条处理函数
# 只按完整 InChIKey 变化删除，不再统计 FIRST_BLOCK
# =========================================================
def process_one_smiles(smi: str, input_inchikey=None, monoisotopic_mass=None):
    base_issue = {
        "smiles_raw": smi,
        "input_inchikey": input_inchikey,
        "input_monoisotopic_mass": monoisotopic_mass,
        "inchi_calc": None,
        "inchikey_calc": None,
        "formula": None,
        "reason": None,
        "parse_error": None,
        "sanitize_error": None,
        "input_inchikey_equal_calc_full": None,
    }

    # 0. 缺失值检查
    if smi is None or smi == "":
        base_issue["reason"] = "MISSING_SMILES"
        return False, None, base_issue

    if input_inchikey is None or input_inchikey == "":
        base_issue["reason"] = "MISSING_INPUT_INCHIKEY"
        return False, None, base_issue

    if monoisotopic_mass is None or str(monoisotopic_mass).strip() == "":
        base_issue["reason"] = "MISSING_MONOISOTOPIC_MASS"
        return False, None, base_issue

    # 1. dot 检查
    if "." in smi:
        base_issue["reason"] = "HAS_DOT"
        return False, None, base_issue

    # 2. parse
    try:
        mol = Chem.MolFromSmiles(smi, sanitize=False)
        if mol is None:
            base_issue["reason"] = "PARSE_FAIL"
            base_issue["parse_error"] = "MolFromSmiles returned None"
            return False, None, base_issue
    except Exception as e:
        base_issue["reason"] = "PARSE_EXCEPTION"
        base_issue["parse_error"] = str(e)
        return False, None, base_issue

    # 3. sanitize
    try:
        Chem.SanitizeMol(mol)
    except Exception as e:
        base_issue["reason"] = "SANITIZE_FAIL"
        base_issue["sanitize_error"] = str(e)
        return False, None, base_issue

    # 4. 分子量（RDKit 计算）
    try:
        exact_mw = Descriptors.ExactMolWt(mol)
    except Exception as e:
        base_issue["reason"] = "EXACT_MW_FAIL"
        base_issue["parse_error"] = str(e)
        return False, None, base_issue

    if exact_mw > MAX_EXACT_MW:
        base_issue["reason"] = "EXACT_MW_GT_1500"
        return False, None, base_issue

    # 5. 逐原子检查
    try:
        unsupported_atoms = set()
        charged_count = 0
        for atom in mol.GetAtoms():
            if atom.GetSymbol() not in FILTER_ATOMS:
                unsupported_atoms.add(atom.GetSymbol())
            if atom.GetFormalCharge() != 0:
                charged_count += 1
    except Exception as e:
        base_issue["reason"] = "ATOM_CHECK_FAIL"
        base_issue["parse_error"] = str(e)
        return False, None, base_issue

    if unsupported_atoms:
        base_issue["reason"] = "UNSUPPORTED_ATOMS"
        return False, None, base_issue

    if charged_count > 0:
        base_issue["reason"] = "ATOM_FORMAL_CHARGE_NONZERO"
        return False, None, base_issue

    # 6. 生成 formula / inchi / inchikey
    try:
        formula = rdMolDescriptors.CalcMolFormula(mol)
        inchi = Chem.MolToInchi(mol)
        inchikey = Chem.InchiToInchiKey(inchi) if inchi else None
    except Exception as e:
        base_issue["reason"] = "INCHI_FAIL"
        base_issue["parse_error"] = str(e)
        return False, None, base_issue

    if not inchikey:
        base_issue["reason"] = "INCHIKEY_EMPTY"
        return False, None, base_issue

    base_issue.update({
        "formula": formula,
        "inchi_calc": inchi,
        "inchikey_calc": inchikey
    })

    # 7. 严格结构比较：只比较完整 InChIKey
    full_equal = int(input_inchikey == inchikey)
    base_issue["input_inchikey_equal_calc_full"] = full_equal

    if full_equal == 0:
        base_issue["reason"] = "STRUCTURE_CHANGED_FULL_ONLY"
        return False, None, base_issue

    # 8. 通过
    clean_record = {
        "inchi": inchi,
        "SMILES": smi,
        "inchikey": inchikey,
        "formula": formula,
        "MONOISOTOPIC_MASS": monoisotopic_mass
    }
    return True, clean_record, None


# =========================================================
# 主程序
# =========================================================
def main():
    seen_inchikey = set()
    summary = {
        "total_rows": 0,
        "cleaned_rows": 0,
        "removed_rows": 0,
        "duplicate_removed_rows": 0,
    }

    with open(CLEAN_FILE, "w", newline="", encoding="utf-8-sig") as f_clean, \
         open(ISSUE_FILE, "w", newline="", encoding="utf-8-sig") as f_issue:

        clean_writer = None
        issue_writer = None

        for file_path in INPUT_FILES:
            print(f"\n正在处理文件: {file_path}")

            for row_index, smi, input_ik, mono_mass in tqdm(
                iter_rows_from_excel(file_path, SMILES_COL, INPUT_INCHIKEY_COL, MONOISOTOPIC_MASS_COL)
            ):
                summary["total_rows"] += 1

                ok, clean_rec, issue_rec = process_one_smiles(smi, input_ik, mono_mass)

                if ok:
                    ik = clean_rec["inchikey"]

                    if ik in seen_inchikey:
                        summary["duplicate_removed_rows"] += 1
                        summary["removed_rows"] += 1

                        issue_row = clean_rec.copy()
                        issue_row.update({
                            "source_file": file_path,
                            "row_index_in_source": row_index,
                            "reason": "DUPLICATE_FULL_INCHIKEY",
                            "parse_error": None,
                            "sanitize_error": None,
                            "input_inchikey": input_ik,
                            "input_inchikey_equal_calc_full": int(input_ik == clean_rec["inchikey"]),
                        })

                        if issue_writer is None:
                            issue_writer = csv.DictWriter(f_issue, fieldnames=list(issue_row.keys()))
                            issue_writer.writeheader()

                        issue_writer.writerow(issue_row)

                    else:
                        seen_inchikey.add(ik)
                        summary["cleaned_rows"] += 1

                        if clean_writer is None:
                            clean_writer = csv.DictWriter(
                                f_clean,
                                fieldnames=["inchi", "SMILES", "inchikey", "formula", "MONOISOTOPIC_MASS"]
                            )
                            clean_writer.writeheader()

                        clean_writer.writerow(clean_rec)

                else:
                    summary["removed_rows"] += 1
                    issue_row = issue_rec.copy()
                    issue_row["source_file"] = file_path
                    issue_row["row_index_in_source"] = row_index

                    if issue_writer is None:
                        issue_writer = csv.DictWriter(f_issue, fieldnames=list(issue_row.keys()))
                        issue_writer.writeheader()

                    issue_writer.writerow(issue_row)

    print("\n清洗完成")
    print(f"清洗后文件: {CLEAN_FILE}")
    print(f"问题文件: {ISSUE_FILE}")
    print("Summary:", summary)
    print(f"unique_full_inchikey_after_cleaning: {len(seen_inchikey)}")


if __name__ == "__main__":
    main()