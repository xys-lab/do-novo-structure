# -*- coding: utf-8 -*-

import csv
from pathlib import Path
from collections import Counter

from openpyxl import load_workbook
from tqdm import tqdm

from rdkit import Chem, RDLogger
from rdkit.Chem import Descriptors, rdMolDescriptors


# =========================================================
# 配置区
# =========================================================
INPUT_FILES = [
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump1_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump2_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump3_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump4_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump5_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump6_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump7_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump8_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump9_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump10_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump11_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump12_filled.xlsx",
    r"D:\实验\数据\DSSTox_Feb_2024\filled_outputs\DSSToxDump13_filled.xlsx",
   # 可以继续添加更多文件
]


SMILES_COL = "SMILES"
INPUT_INCHIKEY_COL = "INCHIKEY"
MONOISOTOPIC_MASS_COL = "MONOISOTOPIC_MASS"

MAX_EXACT_MW = 1500.0
FILTER_ATOMS = {'C', 'N', 'O', 'S', 'F', 'Cl', 'Br', 'I', 'P', 'H'}

OUTPUT_DIR = Path(r"D:\实验\数据\DSSTox_Feb_2024")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

CLEAN_FILE = OUTPUT_DIR / "cleaned_DSSTox_canonical.csv"

# 是否保留立体信息
# True: canonical SMILES 保留立体信息
# False: canonical SMILES 不保留立体信息
USE_ISOMERIC_SMILES = True

# =========================================================
# 静默 RDKit warning
# =========================================================
RDLogger.logger().setLevel(RDLogger.CRITICAL)
# =========================================================
# 工具函数：统一 float 输出为 9 位小数
# =========================================================
def format_float_9(x):
    if x is None:
        return ""
    try:
        return f"{float(x):.9f}"
    except Exception:
        return ""


# =========================================================
# Excel 流式读取
# =========================================================
def iter_rows_from_excel(file_path, smiles_col, inchikey_col, mono_mass_col):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header_clean = [str(x).strip() if x is not None else "" for x in header]

    required_cols = [smiles_col, inchikey_col, mono_mass_col]
    for col in required_cols:
        if col not in header_clean:
            wb.close()
            raise ValueError(f"{file_path} 中找不到必要列: {col}")

    smiles_idx = header_clean.index(smiles_col)
    inchikey_idx = header_clean.index(inchikey_col)
    mono_mass_idx = header_clean.index(mono_mass_col)

    for row_idx, row in enumerate(rows, start=2):
        if row is None:
            continue

        smi = row[smiles_idx] if smiles_idx < len(row) else None
        input_ik = row[inchikey_idx] if inchikey_idx < len(row) else None
        mono_mass = row[mono_mass_idx] if mono_mass_idx < len(row) else None

        smi = str(smi).strip() if smi is not None else None
        input_ik = str(input_ik).strip() if input_ik is not None else None

        yield row_idx, smi, input_ik, mono_mass

    wb.close()


# =========================================================
# 单条处理函数
# 返回:
#   ok: 是否保留
#   clean_record: 保留时输出记录
#   reason: 不保留时的原因
# =========================================================
def process_one_smiles(smi: str, input_inchikey=None, monoisotopic_mass=None):
    # 0. 缺失值检查：只检查 SMILES 和 INCHIKEY
    if smi is None or smi == "":
        return False, None, "MISSING_SMILES"

    if input_inchikey is None or input_inchikey == "":
        return False, None, "MISSING_INPUT_INCHIKEY"

    # 1. dot 检查
    if "." in smi:
        return False, None, "HAS_DOT"

    # 2. parse
    try:
        mol = Chem.MolFromSmiles(smi, sanitize=False)
        if mol is None:
            return False, None, "PARSE_FAIL"
    except Exception:
        return False, None, "PARSE_EXCEPTION"

    # 3. sanitize
    try:
        Chem.SanitizeMol(mol)
    except Exception:
        return False, None, "SANITIZE_FAIL"

    # 4. 逐原子检查
    try:
        unsupported_atoms = set()
        charged_count = 0

        for atom in mol.GetAtoms():
            if atom.GetSymbol() not in FILTER_ATOMS:
                unsupported_atoms.add(atom.GetSymbol())
            if atom.GetFormalCharge() != 0:
                charged_count += 1

    except Exception:
        return False, None, "ATOM_CHECK_FAIL"

    if unsupported_atoms:
        return False, None, "UNSUPPORTED_ATOMS"

    if charged_count > 0:
        return False, None, "ATOM_FORMAL_CHARGE_NONZERO"

    # 5. 基于同一个 sanitize 后的 mol 统一生成信息
    try:
        canonical_smiles = Chem.MolToSmiles(
            mol,
            canonical=True,
            isomericSmiles=USE_ISOMERIC_SMILES
        )
        formula = rdMolDescriptors.CalcMolFormula(mol)
        exact_mass = float(Descriptors.ExactMolWt(mol))
        inchi = Chem.MolToInchi(mol)
        inchikey = Chem.InchiToInchiKey(inchi) if inchi else None
    except Exception:
        return False, None, "PROPERTY_CALC_FAIL"

    # 6. 分子量过滤
    if exact_mass > MAX_EXACT_MW:
        return False, None, "EXACT_MW_GT_1500"

    # 7. InChIKey 是否成功生成
    if not inchikey:
        return False, None, "INCHIKEY_EMPTY"

    # 8. 完整 InChIKey 一致性检查
    # 比较的是：
    #   原始文件中的 INCHIKEY
    #   vs
    #   RDKit 根据原始 SMILES 识别出的结构算出的 InChIKey
    if input_inchikey != inchikey:
        return False, None, "STRUCTURE_CHANGED_FULL_ONLY"

    # 9. 通过
    clean_record = {
        "SMILES": canonical_smiles,
        "formula": formula,
         "exact_mass": format_float_9(exact_mass),
        "inchi": inchi,
        "inchikey": inchikey,
        "MONOISOTOPIC_MASS": monoisotopic_mass,
    }
    return True, clean_record, None


# =========================================================
# 主程序
# =========================================================
def main():
    seen_inchikey = set()

    summary = {
        "total_rows": 0,
        "cleaned_rows": 0,
        "removed_rows": 0,
        "duplicate_removed_rows": 0,
        "full_inchikey_mismatch_removed_rows": 0,
    }

    reason_counter = Counter()

    with open(CLEAN_FILE, "w", newline="", encoding="utf-8-sig") as f_clean:
        clean_writer = csv.DictWriter(
            f_clean,
            fieldnames=[
                "SMILES",
                "formula",
                "exact_mass",
                "inchi",
                "inchikey",
                "MONOISOTOPIC_MASS",
            ]
        )
        clean_writer.writeheader()

        for file_path in INPUT_FILES:
            print(f"\n正在处理文件: {file_path}")

            for row_index, smi, input_ik, mono_mass in tqdm(
                iter_rows_from_excel(file_path, SMILES_COL, INPUT_INCHIKEY_COL, MONOISOTOPIC_MASS_COL)
            ):
                summary["total_rows"] += 1

                ok, clean_rec, reason = process_one_smiles(smi, input_ik, mono_mass)

                if ok:
                    ik = clean_rec["inchikey"]

                    # 去重：完整 InChIKey 去重
                    if ik in seen_inchikey:
                        summary["removed_rows"] += 1
                        summary["duplicate_removed_rows"] += 1
                        reason_counter["DUPLICATE_FULL_INCHIKEY"] += 1
                    else:
                        seen_inchikey.add(ik)
                        summary["cleaned_rows"] += 1
                        clean_writer.writerow(clean_rec)

                else:
                    summary["removed_rows"] += 1
                    reason_counter[reason] += 1

                    if reason == "STRUCTURE_CHANGED_FULL_ONLY":
                        summary["full_inchikey_mismatch_removed_rows"] += 1

    # =====================================================
    # 终端统计输出
    # =====================================================
    print("\n================ 清洗完成 ================")
    print(f"清洗后文件: {CLEAN_FILE}")
    print("Summary:")
    for k, v in summary.items():
        print(f"  {k}: {v}")

    print(f"\nunique_full_inchikey_after_cleaning: {len(seen_inchikey)}")

    print("\n================ 各类删除原因统计 ================")
    if len(reason_counter) == 0:
        print("没有删除记录。")
    else:
        for reason, count in reason_counter.most_common():
            print(f"{reason}: {count}")

    print("\n================ 完整 InChIKey 过滤统计 ================")
    print(f"因完整 InChIKey 不一致被过滤的数据量: {summary['full_inchikey_mismatch_removed_rows']}")

    if summary["total_rows"] > 0:
        ratio_total = summary["full_inchikey_mismatch_removed_rows"] / summary["total_rows"] * 100
        print(f"完整 InChIKey 不一致占总数据比例: {ratio_total:.6f}%")

    if summary["removed_rows"] > 0:
        ratio_removed = summary["full_inchikey_mismatch_removed_rows"] / summary["removed_rows"] * 100
        print(f"完整 InChIKey 不一致占已删除数据比例: {ratio_removed:.6f}%")

    print("======================================================\n")


if __name__ == "__main__":
    main()